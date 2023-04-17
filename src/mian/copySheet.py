# time: 2023/4/16 22:14
# file: copySheet.py

# todo 看看多行注释的使用
# 要使用 openpyxl 复制工作表，您可以按照以下步骤操作：
# 导入 openpyxl 库。
# 打开包含要复制的工作表的 Excel 工作簿。
# 获取要复制的工作表的工作表对象。
# 创建一个新的工作表对象。
# 使用 copy() 方法将旧工作表的内容复制到新工作表。
# 保存 Excel 工作簿。
# 以下是如何使用 openpyxl 复制工作表的示例：
# import pandas as pd
#
# # 读取原始工作表数据
# filename = r"C:\Users\king\Desktop\测试.xlsx"
# sheetname = '所有数据'
# df = pd.read_excel(filename, sheet_name=sheetname)
#
# # 将数据写入新的工作表
# new_sheetname = 'New Sheet'
# # df.to_excel(filename, sheet_name=new_sheetname, index=False)
#
# print(f'Successfully copied {sheetname} to {new_sheetname}')
#
# import pandas as pd
#
# # 打开 Excel 文件并获取所有工作表的名称列表
# filename = r"C:\Users\king\Desktop\0221结果.xlsx"
# xlsx = pd.ExcelFile(filename)
# sheetnames = xlsx.sheet_names
#
# # 遍历所有工作表并读取指定工作表的数据
# for sheetname in sheetnames:
#     df = pd.read_excel(xlsx, sheet_name=sheetname)
#     print(df.head())


# 这个是使用 pd去复制openpyxl
def copy_excel_sheet(filename, table_index=1):
    # 读取数据并创建 DataFrame 对象
    df = pd.ExcelFile(filename)
    # 读取第二个表格
    sheet_names = df.sheet_names
    df = df.parse(sheet_names[table_index])
    # 将 DataFrame 对象复制到工作表中
    # 打开 Excel 文件并获取正在写入的工作表
    wb = openpyxl.load_workbook(filename)
    ws = wb.create_sheet(f"{sheet_names[table_index]}_new")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # 保存文件
    wb.save(filename)


# copy_excel_sheet(r"C:\Users\king\Desktop\测试.xlsx")
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


def copy_excel_sheet_pd(filename, table_index=1):
    # 打开 Excel 文件并获取工作表
    wb = openpyxl.load_workbook(filename)
    # 获取表数据
    ws = wb[wb.sheetnames[table_index]]
    # 关闭文件
    wb.close()
    ws.cell(row=1, column=1).value = "这个是新的测试"
    # 获取工作表中的数据并转换为 DataFrame 对象
    copy_excel_pd(ws)


# 使用pandas复制openpyxl的excel，然后复制进去
def copy_excel_pd(ws, filename="new.xlsx", sheet_name="new_sheet"):
    """
    :param ws: 传入excel的sheet
    :param filename:
    :param sheet_name:
    :return:
    """
    data_rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
    # 使用 DataFrame 对象将数据写入 Excel 文件
    df.to_excel(f"{filename}", sheet_name=f'{sheet_name}', index=False)


copy_excel_sheet_pd(r"C:\Users\king\Desktop\测试.xlsx")
