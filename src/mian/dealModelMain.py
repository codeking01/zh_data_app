import copy
import os
from math import ceil

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score
from sklearn.model_selection import GridSearchCV, RandomizedSearchCV


class AssignDataUtils:
    def __init__(self):
        """ 划分数据的
        """
        pass

    @staticmethod
    def get_train_test_data(x_data=None, y_data=None):
        """ 按照奇偶分
        :param X_data:
        :param Y_data:
        :return: X_train, X_test, y_train, y_test
        """
        # 拷贝处理
        x_data_copy = copy.deepcopy(x_data)
        y_data_copy = copy.deepcopy(y_data)
        # 转化成ndarray操作
        X_train = []
        X_test = []
        Y_train = []
        Y_test = []
        iter_index = 1
        for i in range(x_data_copy.shape[0]):
            # 3 个一组 1 3 训练 2 测试
            if iter_index == 4:
                iter_index = 1
                X_train.append(x_data_copy[i, :])
                Y_train.append(y_data_copy[i])
            elif iter_index == 2:
                X_test.append(x_data_copy[i, :])
                Y_test.append(y_data_copy[i])
            else:
                X_train.append(x_data_copy[i, :])
                Y_train.append(y_data_copy[i])
            iter_index += 1
        X_train = np.array(X_train)
        X_test = np.array(X_test)
        Y_train = np.array(Y_train)
        Y_test = np.array(Y_test)
        return X_train, X_test, Y_train, Y_test

    @staticmethod
    def get_develop_pred_data(excel_data=None, boundary_x=None, boundary_y=None):
        """
        :param excel_data:
        :param boundary_x:
        :param boundary_y:
        :return:
        """
        excel_data_copy = copy.deepcopy(excel_data)
        return excel_data_copy[boundary_x[0]:boundary_x[-1], boundary_y[0]:boundary_y[-1]]


class DataUtils:
    def __init__(self):
        pass

    @staticmethod
    def deal_sorted_y_data(data=None, range=0.1):
        """
        :param data: 需要归一化的元数据
        :param range: 幅度范围 默认0.1
        :return:
        """
        # 用来记录原顺序的数组，取反
        # order = np.argsort(-data)
        sort_data = copy.deepcopy(data)
        sort_data = sorted(sort_data)
        # sorted返回的是数组类型，这个地方需要转成ndarray
        sort_data = np.array(sort_data)
        data_length = sort_data.shape[0]
        # 根据5%为1，中间90%为0，后5%为-1
        # 这个地方要考虑临界条件，当处于临界条件的值，全部分配给两端
        bound_min_value = sort_data[int(ceil(range * data_length))]
        bound_max_value = sort_data[int(ceil((1 - range) * data_length))]
        min_data = np.where(data <= bound_min_value)
        max_data = np.where(data >= bound_max_value)
        normal_data = np.where((data > bound_min_value) & (data < bound_max_value))
        # 统一进行替换为-1，1，0
        data[min_data] = -1
        data[max_data] = 1
        data[normal_data] = 0
        return data, bound_min_value, bound_max_value

    @staticmethod
    # 根据Y_data_boundsMin和Y_data_boundsMax处理Y_data
    def deal_verify_ydata(Verify_Ydata, Y_data_boundsMin, Y_data_boundsMax):
        # 将pred_Ydata 替换成-1，0，1的分类
        min_data = np.where(Verify_Ydata <= Y_data_boundsMin)
        # print('min_data', min_data)
        max_data = np.where(Verify_Ydata >= Y_data_boundsMax)
        # print('max_data', max_data)
        normal_data = np.where((Verify_Ydata > Y_data_boundsMin) & (Verify_Ydata < Y_data_boundsMax))
        # print('normal_data', normal_data)
        # 统一进行替换为-1，1，0
        Verify_Ydata[min_data] = -1
        Verify_Ydata[max_data] = 1
        Verify_Ydata[normal_data] = 0
        return Verify_Ydata

    @staticmethod
    def convert_to_month(x_data=None, col=0):
        """
        :param x_data:
        :return: 将第一列(默认)转化为月份
        """
        x_data_copy = copy.deepcopy(x_data)
        first_row_data = list(x_data_copy[:, col])
        row_length = len(first_row_data)
        try:
            for index in range(row_length):
                if type(first_row_data[index]) == str:
                    # 判断是否存在 “-”
                    if "-" in first_row_data[index]:
                        first_row_data[index] = float(first_row_data[index].split("-")[-1])
                        # 取出最后两位作为月份
                        # first_row_data[index]=float(first_row_data[index][-2:])
        except Exception as e:
            print(e)
        # 取出最后两位作为月份
        # first_row_data = [float(item[-2:]) for item in first_row_data]
        first_row_data = np.array(first_row_data)
        x_data[:, col] = first_row_data
        return x_data


class ModelUtils:
    def __init__(self):
        """ 训练模型的类工具
        """
        pass

    @staticmethod
    def normalization_develop_y(y_data=None):
        """
        :param y_data:建模的Y(源数据，未转化-1 0 1)
        :return:  y_data, y_data_bounds_min, y_data_bounds_max
        """
        # 获取 建模和预测转化好的Y
        y_data, y_data_bounds_min, y_data_bounds_max = DataUtils.deal_sorted_y_data(data=copy.deepcopy(y_data))
        return y_data, y_data_bounds_min, y_data_bounds_max

    @staticmethod
    def normalization_predict_y(y_data=None, bounds_min=None, bounds_max=None):
        """
        :param y_data:
        :param bounds_min: 建模的时候的归一化的边界min
        :param bounds_max:
        :return:
        """
        # 验证（预测）的Y
        data = copy.deepcopy(y_data)
        data = np.array(data, dtype=float)
        return DataUtils.deal_verify_ydata(data, bounds_min, bounds_max)

    @staticmethod
    def develop_save_model(model=None, model_name=None, model_dict=None, save_path="./all_models",
                           x_train=None, y_train=None, x_test=None, y_test=None, train_numbers=None):
        """
        :param x_test:
        :param y_test:
        :param model: 需要训练的模型
        :param model_name:
        :param model_dict:
        :param save_path: 保存的路径，默认是当前文件夹下的 ./models
        :param x_train:
        :param y_train:
        :param train_numbers: 训练次数
        :return:
        """
        CommonUtils.check_dir(save_path)
        # todo 根据 train_numbers决定训练次数，只保留一个最好的模型
        max_accuracy = 0
        for i in range(train_numbers):
            # 只存储一个大模型
            model.fit(x_train, y_train.astype(float))
            # 训练完毕后，计算测试数据的模型的准确率
            test_accuracy = accuracy_score(y_test, model.predict(x_test))
            if test_accuracy > max_accuracy:
                max_accuracy = test_accuracy
                print(f"共{train_numbers}次,当前循环{i}次数,提升了模型准确率:{max_accuracy}")

        model_dict.update({f'{model_name}': model})
        # 将字典中的所有模型保存到一个 joblib。可以考虑最后再保存，这样保证训练和预测可以同时进行
        joblib.dump(model_dict, f'{save_path}/models_temp.joblib')


class OperateExcel:
    def __init__(self):
        """ 操作Excel的
        """
        pass

    @staticmethod
    # 使用pandas复制openpyxl的excel，然后复制进去
    def copy_excel_pd(ws=None, filename="new.xlsx", sheet_name="result_data"):
        """
        :param ws: work_sheet, 传入excel的sheet
        :param filename:
        :param sheet_name:
        :return:
        """
        data_rows = list(ws.iter_rows(values_only=True))
        df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
        # 使用 DataFrame 对象将数据写入 Excel 文件
        # 获取前面的路径
        filename = os.path.dirname(filename) + f"\\预测结果.{filename.split('.')[-1]}"
        # filename=lambda filename:filename.split("\/")[-1]
        df.to_excel(f"{filename}", sheet_name=f'{sheet_name}', index=False)

    @staticmethod
    def record_excel(use_raw_list, excel_col, excel_raw, predict_excel_table, predict_y_item, model_predict_data,
                     accuracy, max_row, save_filename):
        """
        :param save_filename: 保存的路径与名字
        :param max_row: excel最大列号
        :param accuracy: 准确率
        :param use_raw_list: 可以的列
        :param excel_col: excel当前的列号
        :param excel_raw: excel当前的行号
        :param predict_excel_table:  excel当前的sheet
        :param predict_y_item: 原始excel的归一化后的数据
        :param model_predict_data: 预测出的数据
        :return:
        """
        # 写入excel里面
        for i in range(use_raw_list.size):
            predict_excel_table.cell(use_raw_list[i] + excel_raw,
                                     excel_col).value = f"real:{predict_y_item[i]},predict:{model_predict_data[i]}"
        # 最后写上准确率
        predict_excel_table.cell(max_row, excel_col).value = f"{accuracy}"
        OperateExcel.copy_excel_pd(ws=predict_excel_table, filename=f"{save_filename}", sheet_name="new_sheet")


class CleanData:
    def __init__(self):
        """ 清理数据的通用方法
        """
        pass

    @staticmethod
    def del_raw_col_data(data_value, flag):
        """
        :param data_value: 需要处理的数据
        :param flag: flag是1则删除列，如果是0则删除行
        :return: final_data, del_raws/del_cols
        """
        # 记录需要删除的行或者列
        del_list = []
        data_value_copy = copy.deepcopy(data_value)
        # 先拿到元数据的长度
        pd_data_value_copy = pd.DataFrame(copy.deepcopy(data_value_copy))
        # 获取所有存在的空值
        # result_list = np.array(pd_data_value_copy.iloc[:, :].isnull().values.tolist())
        result_list = pd_data_value_copy.iloc[:, :].isnull()
        # 删除缺失值过多的列 直接利用 df.isnull().sum(axis=0))，0是计算列存在的缺失值，1是行
        if flag == 1:
            lock_col_list = result_list.sum(axis=0)
            # 判断缺失值是否大于0.3
            del_list = [i for i in range(lock_col_list.size) if
                        lock_col_list[i] / pd_data_value_copy.iloc[:, 0].size > 0.3]
        elif flag == 0:
            lock_raw_list = result_list.sum(axis=1)
            del_list = [i for i in range(lock_raw_list.size) if lock_raw_list[i] != 0]
        data_value_copy = np.delete(data_value_copy, del_list, axis=flag)
        return data_value_copy, del_list

    @staticmethod
    def record_deletion(del_list=None, use_x_cols=None):
        """
        :param
        del_list: 需要删除的列
        :param
        use_x_cols: 原始列表（包含所有）
        :return:
        """
        # 找到为0的列号
        use_x_cols_zero = np.where(copy.deepcopy(use_x_cols) == 0)[0]
        # 找到需要处理的列
        del_list_temp = use_x_cols_zero[del_list]
        use_x_cols[del_list_temp] = 1

    @staticmethod
    def get_x_y_data(develop_x_data=None, develop_y_item=None):
        """
        :param develop_x_data:
        :param develop_y_item:
        :return:  合并一块删除删除缺失值的行
        """
        develop_x_data_copy = copy.deepcopy(develop_x_data)
        develop_y_item_copy = copy.deepcopy(develop_y_item)
        # 合并数据
        x_y_data = np.c_[develop_x_data_copy, develop_y_item_copy]
        # 删除不可用的行的
        x_y_data = CleanData.del_raw_col_data(data_value=x_y_data, flag=0)
        # 删除缺失行后的x,y
        develop_x_data_copy = x_y_data[0][:, :-1]
        develop_y_item_copy = x_y_data[0][:, -1]
        return develop_x_data_copy, develop_y_item_copy


class CommonUtils:

    def __init__(self):
        pass

    @staticmethod
    def check_contains_letters_or_chinese(arr):
        """ 记录存在汉字和字母的列
        :param arr:
        :return:
        """

        def contains_letters_or_chinese(s):
            for c in s:
                if (u'\u4e00' <= c <= u'\u9fff') or c.isalpha():
                    return True
            return False

        df = pd.DataFrame(copy.deepcopy(arr))
        # mask = df.applymap(lambda x: contains_letters_or_chinese(str(x)))
        mask = df.applymap(lambda x: contains_letters_or_chinese(str(x)) if not pd.isna(x) else False)
        mask = mask.any(axis=0)
        indices = [i for i, x in enumerate(mask.tolist()) if x]
        return indices

    @staticmethod
    def gen_init_data(excel_path=None, excel_sheet=None):
        """
        复制一份去操作
        :param excel_path:
        :param excel_sheet: 如果是传入的字符串，就写名字；或者传入第几个sheet，从0开始
        :return:
        """
        # 初始化
        filename = f'{excel_path}'
        # 复制一份
        init_data = pd.read_excel(io=filename, sheet_name=excel_sheet)
        excel_data = load_workbook(filename)
        load_data = copy.deepcopy(excel_data)
        # 关闭excel
        excel_data.close()
        sheetnames = load_data.sheetnames
        table = load_data[sheetnames[excel_sheet]]
        # 转化为numpy的数据
        init_np_data = np.array(init_data)
        return init_np_data, table, filename, load_data

    @staticmethod
    def start_develop_model(develop_x_data=None, develop_y_dict=None, model_dict=None, use_x_cols=None,
                            train_numbers=None):
        """
        :param train_numbers: 建模的次数
        :param use_x_cols: 建模的时候用到的列号
        :param model_dict:
        :param develop_x_data:
        :param develop_y_dict: 建模的y,是excel列号和numpy数组存储的
        :return:
        """
        # 只操作拷贝数据,挨个遍历key:y_item
        for y_item in develop_y_dict:
            # 每次需要重写拷贝
            # 挨个建模
            develop_y_item = develop_y_dict[y_item]
            develop_x_data_copy, develop_y_item = CleanData.get_x_y_data(develop_x_data, develop_y_item)
            # 归一化，然后记录这个，要和预测的标准保持一致
            develop_y_item, y_data_bounds_min, y_data_bounds_max = ModelUtils.normalization_develop_y(
                y_data=develop_y_item)
            # 这个一块存进字典,前面是y归一化的边界，后面是选取x的列号
            model_dict[f'{y_item}_data'] = [y_data_bounds_min, y_data_bounds_max], use_x_cols

            # 手动划分数据集，各自一半，按照奇偶
            x_train, x_test, y_train, y_test = AssignDataUtils.get_train_test_data(x_data=develop_x_data_copy,
                                                                                   y_data=develop_y_item)
            # rfc_model = RandomForestClassifier(n_estimators=10, criterion='entropy', max_depth=5, min_samples_split=10,
            #                                    min_samples_leaf=5)

            rfc_model = CommonUtils.get_rfc_model(rfc_name="normal", train_numbers=train_numbers)

            ModelUtils.develop_save_model(model=rfc_model, model_name=y_item, model_dict=model_dict,
                                          x_train=x_train, y_train=y_train, x_test=x_test, y_test=y_test,
                                          train_numbers=train_numbers)
        joblib.dump(model_dict, f'./all_models/models.joblib')
        # 删除models_temp.joblib文件
        os.remove(f'./all_models/models_temp.joblib')

    @staticmethod
    def get_rfc_model(rfc_name=None, train_numbers=1):
        """
        :param train_numbers: 默认是1
        :param rfc_name: grid_search(很慢),random_search(较慢),normal(快)
        :return:
        """
        select_rfc_model = ''
        if rfc_name == "grid_search" or (10 < train_numbers < 20):
            param_grid = {'n_estimators': [50, 100, 150],
                          'max_depth': [None, 5, 10],
                          'min_samples_split': [2, 5, 10],
                          'min_samples_leaf': [1, 2, 5],
                          'criterion': ['gini', 'entropy']}
            # 定义随机森林分类器
            clf = RandomForestClassifier(random_state=42)
            # 使用网格搜索进行参数调优
            select_rfc_model = GridSearchCV(clf, param_grid=param_grid, cv=5, n_jobs=-1, scoring='accuracy')
        elif rfc_name == "random_search" or (20 < train_numbers < 100):
            # 定义参数范围
            param_dist = {'n_estimators': list(range(50, 201, 50)),
                          'max_depth': [None] + list(range(5, 16, 5)),
                          'min_samples_split': list(range(2, 11, 2)),
                          'min_samples_leaf': list(range(1, 6)),
                          'criterion': ['gini', 'entropy']}

            # 定义随机森林分类器
            clf = RandomForestClassifier(random_state=42)

            # 使用随机搜索进行参数调优
            select_rfc_model = RandomizedSearchCV(clf, param_distributions=param_dist, n_iter=50, cv=5, n_jobs=-1,
                                                  scoring='accuracy', random_state=42)
        elif rfc_name is None or rfc_name == "normal":
            # 选择最普通的模型
            select_rfc_model = RandomForestClassifier(n_estimators=30, criterion='entropy', max_depth=5,
                                                      min_samples_split=10, min_samples_leaf=5, random_state=42)
        return select_rfc_model

    @staticmethod
    def check_dir(save_path):
        """ 检查目录是否存在，不存在就新建
        :param save_path:
        :return:
        """
        if not os.path.exists(f"{save_path}"):
            os.makedirs(name=f"{save_path}", exist_ok=True)


class OperateModel:
    def __init__(self):
        """ 操作模型的训练预测
        """
        pass

    @staticmethod
    def select_develop_model(excel_path=None, excel_sheet=0, train_numbers=None):
        """
        :param train_numbers: 训练的次数
        :param excel_path: 选择路径
        :param excel_sheet: 0 代表第一张表
        :return: 最后会删除临时文件，并且将其删除，然后保存新的文件
        """
        # 从excel读取数据
        develop_excel_np_data, _, _, _ = CommonUtils.gen_init_data(excel_path=f"{excel_path}",
                                                                   excel_sheet=excel_sheet)
        x_train_boundary = [[2, 21], [27, 30], [34, 49], [57, 70], [83, 87]]
        # 用来训练的y
        y_train_boundary = [[21, 27], [30, 34], [49, 57], [70, 83], [87, 106]]
        # 存储模型的字典
        model_dict = {}
        # 跑每一道工序的循环
        for order_index in range(len(x_train_boundary)):
            # 记录excel 的索引
            develop_y_index_list = [i for i in
                                    range(y_train_boundary[order_index][0], y_train_boundary[order_index][-1])]
            # 记录使用了哪些x列,从2开始
            use_x_cols = np.array([0 for _ in range(x_train_boundary[order_index][-1] - x_train_boundary[0][0])])
            order_x_item = x_train_boundary[order_index]
            order_y_item = y_train_boundary[order_index]
            # 建模数据
            develop_x_data = AssignDataUtils.get_develop_pred_data(excel_data=develop_excel_np_data, boundary_x=[3, -1],
                                                                   boundary_y=[2, order_x_item[-1]])
            # 获取每一道工序的 y
            develop_y_data = AssignDataUtils.get_develop_pred_data(excel_data=develop_excel_np_data, boundary_x=[3, -1],
                                                                   boundary_y=[order_y_item[0], order_y_item[-1]])
            # 转化月份
            develop_x_data = DataUtils.convert_to_month(x_data=develop_x_data)
            # 将建模中的X存在的字母，汉字 记录这个列
            exit_alpha_list = CommonUtils.check_contains_letters_or_chinese(develop_x_data)
            # 对应的列填充1
            CleanData.record_deletion(exit_alpha_list, use_x_cols)
            develop_x_data = np.delete(develop_x_data, exit_alpha_list, 1)
            # 删除缺失值过多的列，并保存del_cols,记录缺失值大于0.3的列
            develop_x_data, del_cols = CleanData.del_raw_col_data(develop_x_data, 1)
            CleanData.record_deletion(del_cols, use_x_cols)
            # 记录能用的列
            use_x_cols = np.where(use_x_cols == 0)[0]
            # develop_x_data = np.delete(develop_x_data, del_cols, axis=1)
            # todo 降维处理,这个地方只从训练集判断相关性，然后统一降维，这个地方存在问题
            # use_able_x_cols = DataDelete(develop_x_data, 0.80)
            # 降维后可用列的X：
            # develop_x_data = develop_x_data[:, use_able_x_cols]
            # 处理建模的y,删除缺失值多的列
            develop_y_data, del_develop_cols = CleanData.del_raw_col_data(develop_y_data, 1)
            # 删除不要的列
            # 计算可以用的列,预测的时候，要对应写进去,并且需要保存当前列的模型
            use_develop_cols = np.delete(develop_y_index_list, del_develop_cols)
            # 存成字典，直接保存去遍历字典
            develop_y_dict = {use_develop_cols[i]: develop_y_data[:, i] for i in range(len(use_develop_cols))}
            print(f"当前内循环{order_y_item[0] + 1},当前excel的索引：{order_y_item[0] + 1}")
            CommonUtils.start_develop_model(develop_x_data=develop_x_data, develop_y_dict=develop_y_dict,
                                            model_dict=model_dict, use_x_cols=use_x_cols, train_numbers=train_numbers)
        print("训练运行结束！")

    @staticmethod
    def select_predict_model(excel_path=None, develop_model_path=r"./all_models/models.joblib", excel_sheet=1):
        """
        :param excel_path:
        :param develop_model_path: 默认的模型路径
        :param excel_sheet: 默认读取第二张表
        :return: 预测好的数据写一个新的excel文件
        """
        # 到时候将这个predict_excel_table复制过去新表
        predict_excel_np_data, predict_excel_table, predict_excel_filename, predict_excel_data = CommonUtils.gen_init_data(
            excel_path=f"{excel_path}", excel_sheet=excel_sheet)
        # 最大行数
        max_row = predict_excel_table.max_row + 1
        # 用来训练的x
        x_predict_boundary = [[2, 21], [27, 30], [34, 49], [57, 70], [83, 87]]
        # 用来训练的y
        y_predict_boundary = [[21, 27], [30, 34], [49, 57], [70, 83], [87, 106]]
        # 获取存储模型的字典
        all_models = joblib.load(f'{develop_model_path}')
        # 跑每一道工序的循环
        for order_index in range(len(x_predict_boundary)):
            # 当前excel行号,其实就是3
            excel_col = y_predict_boundary[order_index][0] + 1
            excel_raw = x_predict_boundary[0][0] + 3
            # 记录excel 的索引
            predict_y_index_list = [i for i in
                                    range(y_predict_boundary[order_index][0], y_predict_boundary[order_index][-1])]
            order_x_item = x_predict_boundary[order_index]
            order_y_item = y_predict_boundary[order_index]

            # 获取预测数据
            predict_x_data = AssignDataUtils.get_develop_pred_data(excel_data=predict_excel_np_data, boundary_x=[3, -1],
                                                                   boundary_y=[2, order_x_item[-1]])
            # 获取每一道工序的 y
            predict_y_data = AssignDataUtils.get_develop_pred_data(excel_data=predict_excel_np_data, boundary_x=[3, -1],
                                                                   boundary_y=[order_y_item[0], order_y_item[-1]])

            # 转化月份
            predict_x_data = DataUtils.convert_to_month(x_data=predict_x_data)
            # 读取模型中的需要的列
            for k in predict_y_index_list:
                if f"{k}_data" in all_models:
                    predict_x_data_copy = copy.deepcopy(predict_x_data)
                    # 获取模型
                    model = all_models[f"{k}"]
                    # 获取归一化数据
                    boundary_list = all_models[f"{k}_data"][0]
                    # 获取可用的X列,和建模的保持一致
                    use_develop_cols = all_models[f"{k}_data"][1]
                    # 获取总的行数，并且生成列表
                    use_raw_list = [i for i in range(len(predict_x_data_copy))]
                    # 复制一份
                    use_raw_list_copy = copy.deepcopy(use_raw_list)
                    # 获取可用的X
                    predict_x_item = predict_x_data_copy[:, use_develop_cols]
                    # 删除缺失值多的行，只需要删除预测的X的行
                    predict_x_item, del_raw_list = CleanData.del_raw_col_data(data_value=predict_x_item, flag=0)
                    # 获取能用的行
                    use_raw_list = np.delete(use_raw_list, del_raw_list)
                    # 获取可用的y列 k-order_y_item[0]是可用的列
                    predict_y_item = predict_y_data[:, k - order_y_item[0]]
                    # 归一化,将内容一块输出，做对比
                    predict_y_item = ModelUtils.normalization_predict_y(predict_y_item, boundary_list[0],
                                                                        boundary_list[1])
                    predict_y_item = predict_y_item[use_raw_list]
                    # print("predict_y_item: ", predict_y_item)
                    # print("predict_y_item [nan]: ",np.sum(np.isnan(predict_y_item)))
                    # print("predict_y_item [!nan]: ", np.sum(np.logical_not(np.isnan(predict_y_item))))
                    model_predict_data = model.predict(predict_x_item)
                    model_predict_data = np.array(model_predict_data, dtype=int)
                    # 记录准确率  要去除空值再去计算准确率 排查最后几个准确率很低的原因 排查excel损坏的原因,这个地方需要排除predict_y_item空值
                    predict_accuracy = np.where([model_predict_data - predict_y_item][0] == 0)[0][:].size / np.sum(
                        np.logical_not(np.isnan(predict_y_item)))
                    print(f"predict_accuracy：{predict_accuracy}")
                    print(f"当前内循环{order_y_item[0] + 1},当前excel的列索引：{k + 1}")
                    OperateExcel.record_excel(use_raw_list=use_raw_list, excel_col=k + 1, excel_raw=excel_raw,
                                              predict_excel_table=predict_excel_table, predict_y_item=predict_y_item,
                                              model_predict_data=model_predict_data, accuracy=predict_accuracy,
                                              max_row=max_row, save_filename=predict_excel_filename)
        print("运行结束！")


if __name__ == '__main__':
    # 这个是建模的函数接口
    # OperateModel.select_develop_model(excel_path=r"C:\Users\king\Desktop\0221结果 - 副本.xlsx")
    # 这个是预测的函数接口
    OperateModel.select_predict_model(excel_path=r"C:\Users\king\Desktop\0221结果 - 副本.xlsx")
