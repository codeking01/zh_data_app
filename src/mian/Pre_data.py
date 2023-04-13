# author: code_king
# time: 2022/2/19 12:37
# file: Pre_data.py
"""
还是5%异常来处理
前面阶段的输入和输出是否都是下一个阶段的输入’和‘前面阶段的输入是下一个阶段的输入’两种方式都尝试一下
"""
import copy
import os
import re
import warnings

import joblib
import numpy as np
import pandas as pd
from numpy import unique
from sklearn.metrics import accuracy_score

from src.utils.model_utils.mian_utils import DataDelete
from src.common_utils.ad_ad_sp import train_model_ad
from src.common_utils.plot_error_distribution import write_images
from src.utils.model_utils.NewPreDeal_Tools import Del_deletion_data, get_final_useablecols, all_ydata, \
    delAndGetCols, get_former_Ydata, convert_to_num, deal_verify_ydata, Deal_sorted_Ydata
from src.utils.plot_utils.build_images.init_build_images import gen_images
from openpyxl import load_workbook

# warnings.filterwarnings(action='always', category=UserWarning)
warnings.filterwarnings(action='ignore')


# base_data_dirs = "../datas"
# 这个pandas处理数据效果不太好 建议用numpy


def gen_init_data(excel_path=None, excel_sheet=None):
    """
    复制一份去操作
    :param excel_path:
    :param excel_sheet:
    :return:
    """
    # 初始化
    excel_data_copy = copy.deepcopy(pd.read_excel(io=f'{excel_path}', sheet_name=excel_sheet))
    init_data = excel_data_copy
    # Preddata = pd.read_excel(io=r'data_original.xlsx', sheet_name='8000D数据标准化-预测用')
    # 用来查看数据结果的excel对象
    filename = f'{excel_path}'
    data = copy.deepcopy(load_workbook(f'{filename}'))
    # sheetnames = data.sheetnames
    # create_sheet
    # if len(sheetnames) < 4:
    #     sheetnames.append('预测结果')
    #     data.create_sheet(sheetnames[-1], len(sheetnames))
    #     # 赋值sheet
    #     # sheet=data[sheetnames[0]]
    #     # content=data.copy_worksheet(sheet)
    #     data.save(f'{filename}')
    # 预测结果的excel表
    sheetnames = data.sheetnames
    table = data[sheetnames[excel_sheet]]

    # 获取数据表对象(建模和与预测的数据)
    init_data = np.array(init_data)
    return init_data, table, filename, data


# 写入excel文件
def write_to_excel(start_index, table, current_column, content):
    # 标签
    for i in range(len(content)):
        table.cell(start_index, current_column).value = content[i]
        start_index += 1


def save_model(model, model_name, X_train, X_test, y_train, y_test, train_numbers):
    """
    :param model: 需要训练的模型
    :param X_train:
    :param X_test:
    :param y_train:
    :param y_test:
    :return:
    """
    # 训练次数
    if not os.path.exists(f"models/{model_name}"):
        os.makedirs(name=f"models/{model_name}", exist_ok=True)
    acc_dic = {}
    for i in range(11):
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        # 取最小的
        model_train_acc = min(model_train_acc, model_acc)
        acc_dic.update({f"model[{i}]": model_train_acc})
        # 模型保存
        joblib.dump(model, f'models/{model_name}/model[{i}].pkl')
    for i in range(11, train_numbers):
        # temp_keys 是键的按照值从小到打排序的列表 [model[1],model[5],...]
        temp_keys = sorted(acc_dic, reverse=True)
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        model_train_acc = min(model_train_acc, model_acc)
        # 替换一个最小的模型
        for key_item_index in range(0, len(temp_keys)):
            if key_item_index != len(temp_keys) - 1:
                if acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc < acc_dic[
                    f"{temp_keys[key_item_index + 1]}"]:
                    acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                    joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')
            elif acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc:
                acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')

    # model1 = joblib.load(filename="filename.pkl")


def get_pred_data(model_name, X_verify_data):
    """
    :param model_name: 模型名字
    :return: pred_data
    """
    # 预测数据
    all_models = os.listdir(f"models/{model_name}")
    model_pred_list = []
    for i in all_models:
        cursor_model = joblib.load(filename=f"models/{model_name}/{i}")
        model_pred_list.append(cursor_model.predict(X_verify_data))
    model_pred_list = np.array(model_pred_list)
    temp_pred_list = []
    # 循环遍历列
    for pred_list_index in range(0, model_pred_list.shape[1]):
        # 预测结果统一,那个数多就等于哪个
        zero_count = np.where(model_pred_list[:, pred_list_index] == 0)[0].shape[0]
        negative_count = np.where(model_pred_list[:, pred_list_index] == -1)[0].shape[0]
        positive_count = np.where(model_pred_list[:, pred_list_index] == 1)[0].shape[0]
        final_dic = {}
        final_dic.update({"0": zero_count})
        final_dic.update({"-1": negative_count})
        final_dic.update({"1": positive_count})
        # 排序，第一个数预测的最多
        final_result = -1
        if final_dic[f"{final_result}"] < final_dic["0"]:
            final_result = 0
        if final_dic[f"{final_result}"] < final_dic["1"]:
            final_result = 1
        temp_pred_list.append(final_result)
    pred_data = np.array(temp_pred_list)
    return pred_data


def get_train_test_acc(model, X_train, y_train, X_test, y_test):
    """
    :param model:
    :param X_train:
    :param y_train:
    :param X_test:
    :param y_test:
    :return: model_train_acc, model_acc
    """
    model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    model_pred = model.predict(X_test)
    model_acc = accuracy_score(y_test.astype(float), model_pred)
    return model_train_acc, model_acc


def get_develop_pred_data(excel_data=None, boundary_x=None, boundary_y=None):
    """
    :param excel_data:
    :param boundary_x:
    :param boundary_y:
    :return:
    """
    excel_data_copy = copy.deepcopy(excel_data)
    return excel_data_copy[boundary_x[0]:boundary_x[-1], boundary_y[0]:boundary_y[-1]]


def get_train_test_data(x_data=None, y_data=None):
    """ 按照奇偶分
    :param X_data:
    :param Y_data:
    :return: X_train, X_test, y_train, y_test
    """
    # 转化成ndarray操作
    X_train = []
    X_test = []
    Y_train = []
    Y_test = []
    iter_index = 1
    for i in range(x_data.shape[0]):
        # 3 个一组 1 3 训练 2 测试
        if iter_index == 4:
            iter_index = 1
            X_train.append(x_data[i, :])
            Y_train.append(y_data[i])
        elif iter_index == 2:
            X_test.append(x_data[i, :])
            Y_test.append(y_data[i])
        else:
            X_train.append(x_data[i, :])
            Y_train.append(y_data[i])
        iter_index += 1
    X_train = np.array(X_train)
    X_test = np.array(X_test)
    Y_train = np.array(Y_train)
    Y_test = np.array(Y_test)
    return X_train, X_test, Y_train, Y_test


def get_usable_data(x_data=None, y_data=None):
    """
    获取删除无效行的数据,并且记录需要删除的行
    :param x_data: 删除无效列的
    :param y_data: 删除无效列的
    :return: 删除无效行以后的数据
    """
    # 先按照列合并
    temp_OriginalXdata = np.column_stack((x_data, y_data))
    # 删除缺失的行
    temp_OriginalXdata, del_raws = Del_deletion_data(temp_OriginalXdata, 0)
    # X_data,Y_data 是处理好(行列均可用)的x,y
    X_data = temp_OriginalXdata[:, :-1]
    Y_data = temp_OriginalXdata[:, -1]
    print(
        f'空列表就是删除干净了.**{list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist()))}', )
    return X_data, Y_data, del_raws


def get_invert_y(y_develop_data=None, y_verify_data=None):
    """
    :param y_develop_data:建模的Y(源数据，未转化-1 0 1)
    :param y_verify_data: 预测的Y(源数据，未转化-1 0 1)
    :return: 转化好的 Y_develop_data,Verify_Ydata
    """
    # 获取 建模和预测转化好的Y
    Y_develop_data_copy = copy.deepcopy(y_develop_data)
    y_develop_data, Y_data_boundsMin, Y_data_boundsMax = Deal_sorted_Ydata(data=Y_develop_data_copy)
    # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
    # 保留原始的数据
    Y_verify_data_copy = copy.deepcopy(y_verify_data)
    # 验证（预测）的Y
    Verify_Ydata = deal_verify_ydata(Y_verify_data_copy, Y_data_boundsMin, Y_data_boundsMax)
    # 查看是否替换成功
    print("查看是否替换成功：", [unique(y_develop_data), Y_data_boundsMin, Y_data_boundsMax],
          np.unique(Verify_Ydata))
    return y_develop_data, Verify_Ydata


def develop_model(x_develop_data=None, x_pred_data=None, y_develop_data=None, y_pred_data=None, final_cols=None,
                  y_filename=None, traverse_index=None, start_month=None,
                  max_progress_row=20, data_interval=10,
                  the_column=None):
    """
    :param the_column: 这个是 out在excel的对应的列号
    :param x_develop_data:
    :param x_pred_data:
    :param y_develop_data:
    :param y_pred_data: 预测的y
    :param final_cols: 可用的列
    :param y_filename: 根据月份的y的文件位置
    :param traverse_index: 遍历次数，可用乘一个固定数，这样让excel的数据加进去
    :param start_month: 开始月份
    :param max_progress_row: 列的列表 最长 默认给了20
    :param data_interval: excel间隔，默认为10
    :return:
    """
    for index in range(len(final_cols)):
        # todo 后面想办法改一下，不要通过序列去传递了，很麻烦的！
        current_index = final_cols[index]
        y_develop_data_curr = y_develop_data[index]
        # y_verify_data_curr = y_pred_data[index]
        # 拷贝原数据，这个画图还是需要使用以前的数据
        Y_develop_data_orinal = copy.deepcopy(y_develop_data_curr)
        # Y_verify_data_orinal = copy.deepcopy(y_verify_data_curr)
        # 已经处理过列了，现在获取行可用的 建模数据,和已经删除的行
        X_develop_data, Y_develop_data, del_raws = get_usable_data(x_data=x_develop_data, y_data=y_develop_data_curr)
        # 处理验证（预测）的X_data,Y_data
        X_verify_data, Y_verify_data, del_raws = get_usable_data(x_data=x_pred_data, y_data=y_verify_data_curr)
        Y_develop_data, Verify_Ydata = get_invert_y(y_develop_data=Y_develop_data, y_verify_data=Y_verify_data)
        # 手动划分数据集，各自一半，按照奇偶
        X_train, X_test, y_train, y_test = get_train_test_data(x_data=X_develop_data, y_data=Y_develop_data)
        # ******************************************************
        # 降维，失败则返回原数据
        # todo 判断是否降维，后面再抽离出来
        """
        try:
            if int(X_train.shape[1]) > 7:
                # 降维
                II = train_model_ad(n_max=max(int(X_train.shape[1]/5),7), cd_n=r"./", cd_e="./", XT_SP_d_train=X_train,
                                    XT_SP_d_test=X_test, Y_train=y_train, Y_test=y_test,
                                    excel_name=f'{start_month + traverse_index}add_result')
                usable_cols = II[-1]
                X_train = X_train[:, usable_cols]
                X_test = X_test[:, usable_cols]
                # ########################################################
                # 验证集
                X_verify_data = X_verify_data[:, usable_cols]
        except Exception as e:
            print(f'降维失败，原因：{e}')
        """

        # ******************************************************
        # 验证集 这个可以不写，只是方便查看
        Y_verify_data = Verify_Ydata
        # 开始训练模型
        from sklearn.ensemble import RandomForestClassifier
        # 设置训练次数
        train_numbers = 15
        # 获取模型  随机森林
        Rfc = RandomForestClassifier()
        model_name = "Rfc"
        save_model(Rfc, model_name, X_train, X_test, y_train, y_test, train_numbers)
        # 随便取一个作为训练集的结果
        current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
        Rfc_train_acc, Rfc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
        Rfc_pred = get_pred_data("Rfc", X_verify_data)
        verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)

        # todo 每个阶段不一样，后面这个会变动
        write_images(data_left=Y_develop_data_orinal, data_right=Y_verify_data_orinal,
                     names=f"{y_filename}/{all_titles[the_column + current_index - 1]}.png", score=verify_y1)
        # 计算准确率
        print('验证集Rfc_acc准确率:', verify_y1)
        #  标签 都加20
        write_to_excel(start_index=5 + traverse_index * data_interval, table=table, current_column=1,
                       content=[f"({start_month + traverse_index}月)训练集个数", "随机森林（Rfc）准确率",
                                "测试集个数", "随机森林（Rfc）准确率",
                                "验证集个数", "随机森林（Rfc）准确率", ])
        # 将训练集写入到excel中
        write_to_excel(start_index=5 + traverse_index * data_interval, table=table, current_column=the_column + index,
                       content=[X_train.shape[0], Rfc_train_acc,
                                X_test.shape[0], Rfc_acc,
                                X_verify_data.shape[0], verify_y1])
        # data_year_list = add_data_year(X_train_num=X_train.shape[0], Rfc_train_acc=Rfc_train_acc,
        #                                X_test_num=X_test.shape[0], Rfc_test_acc=Rfc_acc,
        #                                X_verify_data_num=X_verify_data.shape[0],
        #                                Rfc_verify_acc=verify_y1, data_year_list=data_year_list)
        # max_progress_row 表示当前最长
        # 达到一年会写入 目前是根据 traverse_index达到 12
        write_year_result(the_column=the_column, start_month=start_month, traverse_index=traverse_index, index=index,
                          data_interval=data_interval)
        data.save(filename)


def write_year_result(the_column=None, start_month=None, traverse_index=None, index=None,
                      data_interval=None, max_progress_row=20):
    """
    :param max_progress_row:
    :param the_column:
    :param start_month:
    :param traverse_index:
    :param index:
    :param data_interval:
    :return:
    """
    # 达到一年了，计算一下综合的数据，计算准确率
    if (traverse_index + 1) % 12 == 0:
        # todo 根据excel去操作
        # x_train_index=5 + traverse_index * data_interval
        # x_test_index=5 + traverse_index * data_interval
        # x_verify_index=5 + traverse_index * data_interval
        # [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
        #  x_verify_num_year] = data_year_list
        # x_train_acc_year = x_train_num_right_year / x_train_num_year
        # x_test_acc_year = x_test_num_right_year / x_test_num_year
        # x_verify_acc_year = x_verify_right_year / x_verify_num_year
        write_to_excel(start_index=5 + traverse_index * data_interval + data_interval, table=table, current_column=1,
                       content=[f"(从{start_month}月开始，一年)训练集个数", "随机森林（Rfc）准确率",
                                "测试集个数", "随机森林（Rfc）准确率",
                                "验证集个数", "随机森林（Rfc）准确率", ])
        # 将训练集写入到excel中
        write_to_excel(start_index=5 + traverse_index * data_interval + data_interval, table=table,
                       current_column=the_column + index,
                       content=["x_train_num_year", "x_train_acc_year",
                                "x_test_num_year", "x_test_acc_year",
                                "x_verify_num_year", "x_verify_acc_year"])


def add_data_year(X_train_num=None, Rfc_train_acc=None, X_test_num=None, Rfc_test_acc=None, X_verify_data_num=None,
                  Rfc_verify_acc=None, data_year_list=None):
    """
    这个方法没啥用了
    :param X_train_num:
    :param Rfc_train_acc:
    :param X_test_num:
    :param Rfc_test_acc:
    :param X_verify_data_num:
    :param Rfc_verify_acc:
    :param data_year_list:
    :return:
    """
    # 取出列表数据处理
    [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
     x_verify_num_year] = data_year_list
    # 训练个数
    x_train_num_right_year += X_train_num * Rfc_train_acc
    x_train_num_year += X_train_num
    # 测试
    x_test_num_right_year += X_test_num * Rfc_test_acc
    x_test_num_year += X_test_num
    # 验证
    x_verify_right_year += X_verify_data_num * Rfc_verify_acc
    x_verify_num_year += X_verify_data_num
    # 处理好数据再放回
    temp_list = [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year,
                 x_verify_right_year,
                 x_verify_num_year]
    return temp_list


def write_x_images(develop_x_title_data=None, develop_x_data=None, verify_x_data=None, x_filename="x_images"):
    """
    :param
    develop_x_title_data: 带标题的
    :param
    develop_x_data:
    :return:
    """
    # 将 建模的X 和 预测的X 传入即可，可先将标题带上
    title_data = list(develop_x_title_data)
    for i in range(len(develop_x_title_data)):
        write_images(data_left=develop_x_data[:, i], data_right=verify_x_data[:, i],
                     names=f"{x_filename}/{title_data[i]}.png")


def gen_data_year():
    """ 这个方法没啥用
    :return: data_year 训练，测试，验证的数据
    """
    x_train_num_right_year = 0
    x_train_num_year = 0
    x_test_num_right_year = 0
    x_test_num_year = 0
    x_verify_right_year = 0
    x_verify_num_year = 0
    data_year = [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
                 x_verify_num_year]
    return data_year


def convert_to_month(x_data=None, col=0):
    """
    :param x_data:
    :return: 将第一列(默认)转化为月份
    """
    x_data_copy = copy.deepcopy(x_data)
    first_row_data = list(x_data_copy[:, col])
    row_length = len(first_row_data)
    try:
        for index in range(row_length):
            if type(first_row_data[index]) == str:
                # 判断是否存在 “-”
                if "-" in first_row_data[index]:
                    first_row_data[index] = float(first_row_data[index].split("-")[-1])
                    # 取出最后两位作为月份
                    # first_row_data[index]=float(first_row_data[index][-2:])
    except Exception as e:
        print(e)
    # 取出最后两位作为月份
    # first_row_data = [float(item[-2:]) for item in first_row_data]
    first_row_data = np.array(first_row_data)
    x_data[:, col] = first_row_data
    return x_data


def get_boundary_list(excel_data=None):
    num = 2
    pattern = f"工序{num}-OUT"
    # todo
    pass
    # if re.match(pattern, cur_data):

    return list


def get_month_boundary_list(excel_data=None, start_year=None, start_month=None, pred_count=12, col=2):
    """
    :param excel_data:
    :param start_year:
    :param start_month:
    :param pred_count: 取多少个月的次数，因为中间有月份缺失
    :param col: 从excel第几列开始取数据
    :return:
    """
    start_year = int(start_year)
    start_month = int(start_month)
    # 当前匹配的年月份
    current_date = start_month
    # 是否匹配到开始月份
    start_flag = False
    # 定义正则
    pattern = r"\d{4}-\d{2}"
    # 将一列复制出来，转化为列表，并且转化为字符串
    month_list = copy.deepcopy(excel_data[:, col])
    # 去掉空格，并且转化为字符串
    month_list = [str(i) for i in month_list]
    boundary_list = []
    date_list = []
    for i in range(len(month_list)):
        cur_data = month_list[i]
        if start_flag:
            if pred_count > 0:
                if re.match(pattern, cur_data):
                    if cur_data != current_date:
                        current_date = cur_data
                        # 数据里面的数据有月份缺失
                        boundary_list.append(i)
                        date_list.append(current_date)
                        # 迭代次数减1
                        pred_count = pred_count - 1
            else:
                return boundary_list, date_list
        else:
            # 匹配 xxxx-xx
            if re.match(pattern, cur_data):
                # 匹配到开始月份
                if int(cur_data.split('-')[0]) == start_year and int(cur_data.split('-')[-1]) == start_month:
                    start_flag = True
                    current_date = cur_data
                    # 数据里面的数据有月份缺失
                    boundary_list.append(i)
                    date_list.append(current_date)
                    # 迭代次数减1
                    pred_count = pred_count - 1
    if len(boundary_list) == 0:
        raise Exception(f'{start_year}年的{start_month}月份不存在')
    return boundary_list, date_list


# 预测主函数
def predict_model(develop_file=None,  develop_excel_sheet=0, predict_file=None,predict_excel_sheet=1):
    global develop_init_data, develop_excel_table, develop_excel_filename, develop_excel_data
    develop_init_data, develop_excel_table, develop_excel_filename, develop_excel_data = gen_init_data(
        excel_path=f"{develop_file}", excel_sheet=develop_excel_sheet)
    # todo 这个是预测的文件
    # global predict_init_data, predict_excel_table, predict_excel_filename, predict_excel_data
    # predict_init_data, predict_excel_table, predict_excel_filename, predict_excel_data = gen_init_data(
    #     excel_path=f"{predict_file}", excel_sheet=predict_excel_sheet)

    # 将表的内容 7个数据表 挨个处理
    # 获取源建模的X数据 传入的对象是获取的excel对象
    # 获取标题内容
    global all_titles, process_title
    all_titles = develop_init_data[1, :]
    process_title = develop_init_data[1, 3:]
    # 获取边界的数字
    # month_boundary_list, date_list = get_month_boundary_list(excel_data=init_data, start_year=start_year,
    #                                                          start_month=start_month)

    # 根据月份边际去遍历
    # x_boundary_bottom = [173, 204, 224, 248, 273, 296, 320, 343, 370, 393, 421, 443, 455]
    # x_boundary_bottom = copy.deepcopy(month_boundary_list)
    x_train_boundary = [[2, 21], [27, 30], [34, 49], [57, 70], [83, 87]]
    # 用来验证的数据
    y_verify_boundary = [[21, 27], [30, 34], [49, 57], [70, 83], [87, 106]]
    # x_boundary_left = [320, 343, 370, 393, 421, 443, 455]
    index = 0
    # 开始月份
    # start_month = 3
    # todo 跑每一道工序的循环
    for order_index in range(len(x_train_boundary)):
        order_x_item = x_train_boundary[order_index]
        order_y_item = y_verify_boundary[order_index]
        # 建模数据
        develop_x_data = get_develop_pred_data(excel_data=develop_init_data,
                                               boundary_x=[3, -1],
                                               boundary_y=[2, order_x_item[-1]])
        develop_x_title_data = get_develop_pred_data(excel_data=develop_init_data,
                                                     boundary_x=[1, -1],
                                                     boundary_y=[2, order_x_item[-1]])
        # 获取每一道工序的 y
        develop_y_data = get_develop_pred_data(excel_data=develop_init_data,
                                               boundary_x=[3, -1],
                                               boundary_y=[order_y_item[0], order_y_item[-1]])

        # todo 从另外一个excel获取内容,预测数据
        # pred_x_data = get_develop_pred_data(excel_data=init_data, boundary_x=[x_boundary_bottom[boundary_index],
        #                                                                       x_boundary_bottom[
        #                                                                           boundary_index + 1]],
        #                                     boundary_y=[2, order_x_item[-1]])
        # pred_y_data = get_develop_pred_data(excel_data=init_data, boundary_x=[x_boundary_bottom[boundary_index],
        #                                                                       x_boundary_bottom[
        #                                                                           boundary_index + 1]],
        #                                     boundary_y=[order_y_item[0], order_y_item[-1]])
        # todo
        # 统一关键字

        # 需要将第一列转化为月份
        develop_x_data = convert_to_month(x_data=develop_x_data)
        # pred_x_data = convert_to_month(x_data=pred_x_data)
        # the_predict_former_data, _, del_list = convert_to_num(develop_x_data)
        # todo 将建模中的X存在的字母，汉字删除
        copy_data=copy.deepcopy(develop_x_data)


        develop_x_title_data = np.delete(develop_x_title_data, del_list, axis=1)
        # 删除缺失值过多的列，并保存del_cols
        base_Xdata, del_cols = delAndGetCols(the_predict_former_data)
        # 用del_cols删除验证集X不需要的列
        Verify_Xdata = np.delete(the_Verify_TABLE_data, del_cols, axis=1)
        develop_x_title_data = np.delete(develop_x_title_data, del_cols, axis=1)
        # 降维处理,这个地方只从训练集判断相关性，然后统一降维·
        use_able_x_cols = DataDelete(base_Xdata, 0.80)
        # 所用可用列的X： base_Xdata
        base_Xdata = base_Xdata[:, use_able_x_cols]
        # 获取实际可用列的（行会和Y的数据一起删除） 预测的X数据 Verify_Xdata
        Verify_Xdata = Verify_Xdata[:, use_able_x_cols]
        # 获取可用的列
        develop_x_title_data = develop_x_title_data[0, use_able_x_cols]
        # 处理预测
        final_cols = get_final_useablecols(Original_Yvalue=develop_y_data)
        # 获取可以用“列”的Y_data
        Original_YdataList, _ = all_ydata(final_cols=final_cols, original_yvalue=develop_y_data)
        print(f"当前内循环：{order_y_item[0] + 1}")
        # todo 降维以后在画图
        write_x_images(develop_x_title_data=develop_x_title_data, develop_x_data=base_Xdata,
                       verify_x_data=Verify_Xdata,
                       x_filename=f"x_images[{start_month + index}month_{order_x_item[-1]}col]")
        # 依次遍历每一个Y

        the_column_index = order_y_item[0] + 1
        # todo 传递必须复制一份去操作 y_pred_data
        develop_model(x_develop_data=base_Xdata, x_pred_data=Verify_Xdata,
                      y_develop_data=Original_YdataList,
                      final_cols=final_cols,
                      y_filename=f"y_images[{start_month + index}month_{order_x_item[-1]}col]",
                      traverse_index=index,
                      start_month=start_month, the_column=the_column_index)
    index += 1
    print("运行结束！")

if __name__ =="__mian__":
    pass

