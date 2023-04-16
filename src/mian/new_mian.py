import copy
import os

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_STRING
from openpyxl.workbook import Workbook
from sklearn.ensemble import RandomForestClassifier

from src.utils.model_utils.mian_utils import get_develop_pred_data, convert_to_month, get_train_test_data, \
    develop_save_model, normalization_develop_y, normalization_predict_y


# 使用pandas复制openpyxl的excel，然后复制进去
def copy_excel_pd(ws=None, filename="new.xlsx", sheet_name="new_sheet"):
    """
    :param ws: work_sheet, 传入excel的sheet
    :param filename:
    :param sheet_name:
    :return:
    """
    data_rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
    # 使用 DataFrame 对象将数据写入 Excel 文件
    df.to_excel(f"{filename}", sheet_name=f'{sheet_name}', index=False)


def del_raw_col_data(data_value, flag):
    """
    :param data_value: 需要处理的数据
    :param flag: flag是1则删除列，如果是0则删除行
    :return: final_data, del_raws/del_cols
    """
    # 记录需要删除的行或者列
    del_list = []
    data_value_copy = copy.deepcopy(data_value)
    # 先拿到元数据的长度
    pd_data_value_copy = pd.DataFrame(copy.deepcopy(data_value_copy))
    # 获取所有存在的空值
    # result_list = np.array(pd_data_value_copy.iloc[:, :].isnull().values.tolist())
    result_list = pd_data_value_copy.iloc[:, :].isnull()
    # 删除缺失值过多的列 直接利用 df.isnull().sum(axis=0))，0是计算列存在的缺失值，1是行
    if flag == 1:
        lock_col_list = result_list.sum(axis=0)
        # 判断缺失值是否大于0.3
        del_list = [i for i in range(lock_col_list.size) if lock_col_list[i] / pd_data_value_copy.iloc[:, 0].size > 0.3]
    elif flag == 0:
        lock_raw_list = result_list.sum(axis=1)
        del_list = [i for i in range(lock_raw_list.size) if lock_raw_list[i] != 0]
    data_value_copy = np.delete(data_value_copy, del_list, axis=flag)
    return data_value_copy, del_list

    # 删除缺失值过多的列  # if flag == 1:  #     # 计算多少列  #     cols_length = result_list[0, :].size  #     # 计算里面缺失值大于0.3的  #     for index in range(cols_length):  #         item = result_list[:, index]  #         lock_list = [i for i in item if i]  #         if len(lock_list) / len(item) > 0.3:  #             del_list.append(index)  # # 删除含有缺失值的行  # elif flag == 0:  #     rows_length = result_list[:, 0].size  #     for index in range(rows_length):  #         item = result_list[index, :]  #         # 判断是否有缺失  #         if True in item:  #             del_list.append(index)  # data_value_copy = np.delete(data_value_copy, del_list, axis=flag)  # return data_value_copy, del_list


def gen_init_data(excel_path=None, excel_sheet=None):
    """
    复制一份去操作
    :param excel_path:
    :param excel_sheet: 如果是传入的字符串，就写名字；或者传入第几个sheet，从0开始
    :return:
    """
    # 初始化
    filename = f'{excel_path}'
    # 复制一份
    init_data = pd.read_excel(io=filename, sheet_name=excel_sheet)
    excel_data = load_workbook(filename)
    load_data = copy.deepcopy(excel_data)
    # 关闭excel
    excel_data.close()
    sheetnames = load_data.sheetnames
    table = load_data[sheetnames[excel_sheet]]
    # 转化为numpy的数据
    init_np_data = np.array(init_data)
    return init_np_data, table, filename, load_data


def check_contains_letters_or_chinese(arr):
    """
     记录存在汉字和字母的列
    :param arr:
    :return:
    """

    def contains_letters_or_chinese(s):
        for c in s:
            if (u'\u4e00' <= c <= u'\u9fff') or c.isalpha():
                return True
        return False

    df = pd.DataFrame(copy.deepcopy(arr))
    # mask = df.applymap(lambda x: contains_letters_or_chinese(str(x)))
    mask = df.applymap(lambda x: contains_letters_or_chinese(str(x)) if not pd.isna(x) else False)
    mask = mask.any(axis=0)
    indices = [i for i, x in enumerate(mask.tolist()) if x]
    return indices


def start_develop_model(develop_x_data=None, develop_y_dict=None, model_dict=None, use_x_cols=None):
    """
    :param use_x_cols: 建模的时候用到的列号
    :param model_dict:
    :param develop_x_data:
    :param develop_y_dict: 建模的y,是excel列号和numpy数组存储的
    :return:
    """
    # 只操作拷贝数据,挨个遍历key:y_item
    for y_item in develop_y_dict:
        # 每次需要重写拷贝
        # 挨个建模
        develop_y_item = develop_y_dict[y_item]
        develop_x_data_copy, develop_y_item = get_x_y_data(develop_x_data, develop_y_item)
        # 归一化，然后记录这个，要和预测的标准保持一致
        develop_y_item, y_data_bounds_min, y_data_bounds_max = normalization_develop_y(y_data=develop_y_item)
        # 这个一块存进字典,前面是y归一化的边界，后面是选取x的列号
        model_dict[f'{y_item}_data'] = [y_data_bounds_min, y_data_bounds_max], use_x_cols

        # 手动划分数据集，各自一半，按照奇偶
        x_train, _, y_train, _ = get_train_test_data(x_data=develop_x_data_copy, y_data=develop_y_item)
        develop_save_model(model=RandomForestClassifier(), model_name=y_item, model_dict=model_dict, x_train=x_train,
                           y_train=y_train)
    joblib.dump(model_dict, f'./all_models/models.joblib')
    # 删除models_temp.joblib文件
    os.remove(f'./all_models/models_temp.joblib')


def get_x_y_data(develop_x_data=None, develop_y_item=None):
    """
    :param develop_x_data:
    :param develop_y_item:
    :return:  合并一块删除删除缺失值的行
    """
    develop_x_data_copy = copy.deepcopy(develop_x_data)
    develop_y_item_copy = copy.deepcopy(develop_y_item)
    # 合并数据
    x_y_data = np.c_[develop_x_data_copy, develop_y_item_copy]
    # 删除不可用的行的
    x_y_data = del_raw_col_data(data_value=x_y_data, flag=0)
    # 删除缺失行后的x,y
    develop_x_data_copy = x_y_data[0][:, :-1]
    develop_y_item_copy = x_y_data[0][:, -1]
    return develop_x_data_copy, develop_y_item_copy


def select_develop_model(excel_path=None, excel_sheet=0):
    """
    :param excel_path: 选择路径
    :param excel_sheet: 0 代表第一张表
    :return: 最后会删除临时文件，并且将其删除，然后保存新的文件
    """
    # 从excel读取数据
    develop_excel_np_data, _, _, _ = gen_init_data(excel_path=f"{excel_path}", excel_sheet=excel_sheet)
    # 存储标题
    # global all_titles
    # all_titles = develop_excel_np_data[1, :]
    # 用来训练的x
    x_train_boundary = [[2, 21], [27, 30], [34, 49], [57, 70], [83, 87]]
    # 用来训练的y
    y_train_boundary = [[21, 27], [30, 34], [49, 57], [70, 83], [87, 106]]
    # 存储模型的字典
    model_dict = {}
    # 跑每一道工序的循环
    for order_index in range(len(x_train_boundary)):
        # 记录excel 的索引
        develop_y_index_list = [i for i in range(y_train_boundary[order_index][0], y_train_boundary[order_index][-1])]
        # 记录使用了哪些x列,从2开始
        use_x_cols = np.array([0 for _ in range(x_train_boundary[order_index][-1] - x_train_boundary[0][0])])
        order_x_item = x_train_boundary[order_index]
        order_y_item = y_train_boundary[order_index]
        # 建模数据
        develop_x_data = get_develop_pred_data(excel_data=develop_excel_np_data, boundary_x=[3, -1],
                                               boundary_y=[2, order_x_item[-1]])
        # 获取每一道工序的 y
        develop_y_data = get_develop_pred_data(excel_data=develop_excel_np_data, boundary_x=[3, -1],
                                               boundary_y=[order_y_item[0], order_y_item[-1]])
        # 转化月份
        develop_x_data = convert_to_month(x_data=develop_x_data)
        # 将建模中的X存在的字母，汉字 记录这个列
        exit_alpha_list = check_contains_letters_or_chinese(develop_x_data)
        # 对应的列填充1
        record_deletion(exit_alpha_list, use_x_cols)
        develop_x_data = np.delete(develop_x_data, exit_alpha_list, 1)
        # 删除缺失值过多的列，并保存del_cols,记录缺失值大于0.3的列
        develop_x_data, del_cols = del_raw_col_data(develop_x_data, 1)
        record_deletion(del_cols, use_x_cols)
        # 记录能用的列
        use_x_cols = np.where(use_x_cols == 0)[0]
        # develop_x_data = np.delete(develop_x_data, del_cols, axis=1)
        # todo 降维处理,这个地方只从训练集判断相关性，然后统一降维，这个地方存在问题
        # use_able_x_cols = DataDelete(develop_x_data, 0.80)
        # 降维后可用列的X：
        # develop_x_data = develop_x_data[:, use_able_x_cols]
        # 处理建模的y,删除缺失值多的列
        develop_y_data, del_develop_cols = del_raw_col_data(develop_y_data, 1)
        # 删除不要的列
        # 计算可以用的列,预测的时候，要对应写进去,并且需要保存当前列的模型
        use_develop_cols = np.delete(develop_y_index_list, del_develop_cols)
        # 存成字典，直接保存去遍历字典
        develop_y_dict = {use_develop_cols[i]: develop_y_data[:, i] for i in range(len(use_develop_cols))}
        print(f"当前内循环{order_y_item[0] + 1},当前excel的索引：{order_y_item[0] + 1}")
        start_develop_model(develop_x_data=develop_x_data, use_x_cols=use_x_cols, develop_y_dict=develop_y_dict,
                            model_dict=model_dict)
    print("运行结束！")


def select_predict_model(excel_path=None, develop_model_path=r"./all_models/models.joblib", excel_sheet=1):
    """
    :param
    excel_path:
    :param
    develop_model_path: 默认的模型路径
    :param
    excel_sheet: 默认读取第二张表
    :return: 预测好的数据写一个新的excel文件
    """
    # todo 到时候将这个predict_excel_table复制过去新表
    predict_excel_np_data, predict_excel_table, predict_excel_filename, predict_excel_data = gen_init_data(
        excel_path=f"{excel_path}", excel_sheet=excel_sheet)
    # 最大行数
    max_row = predict_excel_table.max_row + 1
    # 用来训练的x
    x_predict_boundary = [[2, 21], [27, 30], [34, 49], [57, 70], [83, 87]]
    # 用来训练的y
    y_predict_boundary = [[21, 27], [30, 34], [49, 57], [70, 83], [87, 106]]
    # 获取存储模型的字典
    all_models = joblib.load(f'{develop_model_path}')
    # 跑每一道工序的循环
    for order_index in range(len(x_predict_boundary)):
        # 当前excel行号,其实就是3
        excel_col = y_predict_boundary[order_index][0] + 1
        excel_raw = x_predict_boundary[0][0] + 3
        # 记录excel 的索引
        predict_y_index_list = [i for i in
                                range(y_predict_boundary[order_index][0], y_predict_boundary[order_index][-1])]
        order_x_item = x_predict_boundary[order_index]
        order_y_item = y_predict_boundary[order_index]

        # 获取预测数据
        predict_x_data = get_develop_pred_data(excel_data=predict_excel_np_data, boundary_x=[3, -1],
                                               boundary_y=[2, order_x_item[-1]])
        # 获取每一道工序的 y
        predict_y_data = get_develop_pred_data(excel_data=predict_excel_np_data, boundary_x=[3, -1],
                                               boundary_y=[order_y_item[0], order_y_item[-1]])

        # 转化月份
        predict_x_data = convert_to_month(x_data=predict_x_data)
        # TODO 读取模型中的需要的列
        for k in predict_y_index_list:
            if f"{k}_data" in all_models:
                predict_x_data_copy = copy.deepcopy(predict_x_data)
                # 获取模型
                model = all_models[f"{k}"]
                # 获取归一化数据
                boundary_list = all_models[f"{k}_data"][0]
                # 获取可用的X列,和建模的保持一致
                use_develop_cols = all_models[f"{k}_data"][1]
                # 获取总的行数，并且生成列表
                use_raw_list = [i for i in range(len(predict_x_data_copy))]
                # 复制一份
                use_raw_list_copy = copy.deepcopy(use_raw_list)
                # 获取可用的X
                predict_x_item = predict_x_data_copy[:, use_develop_cols]
                # 删除缺失值多的行，只需要删除预测的X的行
                predict_x_item, del_raw_list = del_raw_col_data(data_value=predict_x_item, flag=0)
                # 获取能用的行
                use_raw_list = np.delete(use_raw_list, del_raw_list)
                # 获取可用的y列 k-order_y_item[0]是可用的列
                predict_y_item = predict_y_data[:, k - order_y_item[0]]
                # 归一化,将内容一块输出，做对比
                predict_y_item = normalization_predict_y(predict_y_item, boundary_list[0], boundary_list[1])
                predict_y_item = predict_y_item[use_raw_list]
                model_predict_data = model.predict(predict_x_item)
                model_predict_data = np.array(model_predict_data, dtype=int)
                # 记录准确率 todo 要去除空值再去计算准确率 排查最后几个准确率很低的原因 排查excel损坏的原因
                predict_accuracy = np.where([model_predict_data - predict_y_item][0] == 0)[0][:].size / predict_y_item[
                                                                                                        :].size
                print(f"predict_accuracy：{predict_accuracy}")
                print(f"当前内循环{order_y_item[0] + 1},当前excel的列索引：{k + 1}")
                record_excel(use_raw_list, k + 1, excel_raw, predict_excel_table, predict_y_item, model_predict_data,
                             predict_accuracy, predict_excel_filename, predict_excel_data,
                             max_row)  # todo 保存到excel中  # if "_" in predict_excel_filename:
                #     predict_excel_data.save(f"{predict_excel_filename.split('.')[-2].split('_')[-2]}_new.{predict_excel_filename.split('.')[-1]}")
                # else:
                # predict_excel_data.save(f"{predict_excel_filename}")
    print("运行结束！")


def record_excel(use_raw_list, excel_col, excel_raw, predict_excel_table, predict_y_item, model_predict_data, accuracy,
                 filename, excel_data, max_row):
    """
    :param filename:
    :param excel_data:
    :param max_row:
    :param accuracy:
    :param use_raw_list: 可以的列
    :param excel_col:
    :param excel_raw:
    :param predict_excel_table:
    :param predict_y_item:
    :param model_predict_data:
    :return:
    """
    # 写入excel里面
    for i in range(use_raw_list.size):
        raw_item = use_raw_list[i]
        predict_excel_table.cell(raw_item + excel_raw, excel_col).data_type = TYPE_STRING
        predict_excel_table.cell(raw_item + excel_raw,
                                 excel_col).value = f"real:{predict_y_item[i]},predict:{model_predict_data[i]}"  # print(predict_excel_table.cell(raw_item + excel_raw, excel_col).value)
    # 最后写上准确率
    predict_excel_table.cell(max_row, excel_col).value = f"{accuracy}"
    copy_excel_pd(ws=predict_excel_table, filename="new.xlsx", sheet_name="new_sheet")
    # excel_data.save(filename)


def record_deletion(del_list=None, use_x_cols=None):
    """
    :param
    del_list: 需要删除的列
    :param
    use_x_cols: 原始列表（包含所有）
    :return:
    """
    # 找到为0的列号
    use_x_cols_zero = np.where(copy.deepcopy(use_x_cols) == 0)[0]
    # 找到需要处理的列
    del_list_temp = use_x_cols_zero[del_list]
    use_x_cols[del_list_temp] = 1


if __name__ == '__main__':
    # 这个是建模的函数接口
    # select_develop_model(excel_path=r"C:\Users\king\Desktop\0221结果 - 副本.xlsx")
    # 这个是预测的函数接口
    select_predict_model(excel_path=r"C:\Users\king\Desktop\0221结果 - 副本.xlsx")
