# author: code_king
# time: 2022/12/1 14:13
# file: init_build_images.py
# warnings.filterwarnings(action='always', category=UserWarning)
import copy
import warnings
from copy import deepcopy

import numpy as np
import pandas as pd

from src.common_utils.plot_error_distribution import write_images
from src.utils.model_utils.NewPreDeal_Tools import check_string

warnings.filterwarnings(action='ignore')


def get_excel_data(excel_name=None):
    # 这个pandas处理数据效果不太好 建议用numpy
    thedata = pd.read_excel(io=f'{excel_name}', sheet_name='8000D数据标准化-建模用')
    Preddata = pd.read_excel(io=f'{excel_name}', sheet_name='8000D数据标准化-预测用')
    # 用来查看数据结果的excel对象
    from openpyxl import load_workbook
    filename = f'{excel_name}'
    data = load_workbook(f'{filename}')
    sheetnames = data.sheetnames
    if len(sheetnames) < 3:
        sheetnames.append('预测结果')
        data.create_sheet(sheetnames[2], len(sheetnames))
        # 赋值sheet
        # sheet=data[sheetnames[0]]
        # content=data.copy_worksheet(sheet)
        data.save(f'{filename}')
    # 预测结果的excel表
    # table = data[sheetnames[2]]
    sheetnames = data.sheetnames
    table = data[sheetnames[2]]

    # 获取数据表对象(建模和与预测的数据)
    thedata = np.array(thedata)
    Preddata = np.array(Preddata)
    return thedata, Preddata, table


# 处理带关键字的
def find_keyword(func):
    def wrapper(*args):
        # 获取列数
        all_cols = args[0].shape[1]
        # del_list 存储需要删除的列
        del_list = []
        for i in range(0, all_cols):
            current_original_col = args[0][:, i]
            # 判断字符串是否存在
            flag = check_string(data=current_original_col)
            # 转化类型
            if flag == True:
                current_pred_col = args[1][:, i]
                args[0][:, i], args[1][:, i], del_list = func(current_original_col.astype(str),
                                                              current_pred_col.astype(str), del_list, i)
        # args0 = np.delete(args[0], del_list, axis=1)
        # args1 = np.delete(args[1], del_list, axis=1)
        # 这个地方不删除，改成全部换成0
        args[0][:, del_list] = 0
        args[1][:, del_list] = 0
        return (args[0], args[1])

    return wrapper


@find_keyword
def convert_to_num(original_data=None, pred_data=None, del_list=None, index=None):
    """
    :param original_data: 建模数据
    :param pred_data: 预测数据
    :param del_list: 需要删除的列
    :param index: 当前索引
    :return: 将关键字转化成数字的两个结果,del_list删除的列
    """
    # 取出每一列的唯一值
    unique_original_value = np.unique(original_data)
    unique_pred_value = np.unique(pred_data)
    # 取出并集,这个地方只需要判断是否关键字一致
    intersection = set(unique_original_value) & set(unique_pred_value)
    if len(intersection) == len(unique_original_value):
        # 如果 'nan'在里面,需要删除掉
        if 'nan' in intersection:
            intersection.remove('nan')
        # 这种才处理
        index = -1
        # 把关键字挨个处理
        for i in intersection:
            original_data = np.where(original_data == f"{i}", index, original_data)
            pred_data = np.where(pred_data == f"{i}", index, pred_data)
            index += 1
    else:
        # 删除这一列
        del_list.append(index)
    return original_data, pred_data, del_list


def get_x_build(deal_data):
    """
    :param deal_data:
    :return:
    """
    deal_data_copy = copy.deepcopy(deal_data)
    deal_data_copy[:, 0:4] = 0
    deal_data_copy[:, 21:27] = 0
    deal_data_copy[:, 30:34] = 0
    deal_data_copy[:, 49:57] = 0
    deal_data_copy[:, 70:83] = 0
    deal_data_copy[:, 70:83] = 0
    deal_data_copy[:, 87:106] = 0
    return deal_data_copy[3:, :]


def Del_deletion_data(dataValue, flag):
    """
    :param dataValue: 需要处理的数据
    :param flag: flag是1则删除列，如果是0则删除行
    :return:
    """
    dataValueCopy = deepcopy(dataValue)
    # 删除缺失值过多的列
    if flag == 1:
        del_cols = []
        # 先拿到元数据的长度
        baseLength = dataValueCopy.shape[0]
        # 这个 用来保存需要删除的列
        iterLength = dataValueCopy.shape[1]
        for i in range(iterLength):
            temp = []
            temp = np.where(np.isnan(dataValueCopy[:, i].astype(float)) == True)[0].tolist() + temp
            # print(temp)
            # 用集合去重
            c = set(temp)
            # 计算缺失的总行数
            lack_of_rows = len(c)
            # 计算一列的数全部为0的也删除，这个地方后期得改成某个数
            if lack_of_rows > 0.3 * baseLength:
                del_cols.append(i)
                # print(dataValueCopy)
        # flag是1则删除列，如果是0则删除行
        # final_data = np.delete(dataValueCopy, del_cols, axis=flag)

        # 这个地方不删除，只做替换为0
        dataValueCopy[:, del_cols] = 0
        final_data = dataValueCopy
        return final_data, del_cols
    # 删除含有缺失值的行
    elif flag == 0:
        # 筛选出含有缺失值的行，用set去重
        del_raws = set(np.where(np.isnan(dataValueCopy.astype(float)) == True)[0].tolist())
        del_raws = list(del_raws)
        final_data = np.delete(dataValueCopy, del_raws, axis=flag)
        return final_data


# 删除多余的列，并且返回处理好的数据已经需要删除的列
def delAndGetCols(deal_value):
    # 删除部分的缺失值过多的列
    temp_TwoToThree_Xdata = Del_deletion_data(deal_value, 1)
    base_temp_TwoToThree_Xdata = temp_TwoToThree_Xdata[0]
    # 获取需要删除的列
    del_cols = temp_TwoToThree_Xdata[1]
    return base_temp_TwoToThree_Xdata, del_cols


def gen_images(iter_left_data=None, iter_right_verify=None, files_name=None,titles=None):
    # 按照列去遍历
    for i in range(iter_left_data.shape[1]):
        if np.where(iter_left_data[:, i] == 0)[0].shape[0] == iter_left_data[:, i].shape[0] or \
                np.where(iter_right_verify[:, i] == 0)[0].shape[0] == iter_right_verify[:, i].shape[0]:
            continue
        else:
            write_images(data_left=iter_left_data[:, i], data_right=iter_right_verify[:, i],names=f"{files_name}/{titles[i]  }.png")




if __name__ == '__main__':
    thedata, Preddata, table = get_excel_data(excel_name="data_original.xlsx")
    all_titles=thedata[1,:]
    # 获取所有的X,将除了X的外的数字全部变成0
    x_build = get_x_build(deal_data=thedata)
    x_pred = get_x_build(deal_data=Preddata)
    # 统一关键字
    the_Former_data, the_Verify_TABLE_data = convert_to_num(x_build, x_pred)
    # 替换缺失值过多的列为0，并保存del_cols
    base_Xdata, del_cols = delAndGetCols(the_Former_data)
    # 用del_cols 替换验证集X不需要的列
    # Verify_Xdata = np.delete(the_Verify_TABLE_data, del_cols, axis=1)
    the_Verify_TABLE_data[:, del_cols] = 0
    Verify_Xdata = the_Verify_TABLE_data
    base_Xdata = Del_deletion_data(base_Xdata, 0)
    Verify_Xdata = Del_deletion_data(Verify_Xdata, 0)
    # todo
    gen_images(iter_left_data=base_Xdata, iter_right_verify=Verify_Xdata, files_name="x_images", titles=all_titles)
    print("X_画图 结束！！")
    # 依次获取 X的工序的值

