# author: code_king
# time: 2022/2/19 12:37
# file: Pre_data.py
"""
还是5%异常来处理
前面阶段的输入和输出是否都是下一个阶段的输入’和‘前面阶段的输入是下一个阶段的输入’两种方式都尝试一下
"""
import copy
import math
import os
import re
import warnings

import joblib
import numpy as np
import pandas as pd
from numpy import unique
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score

from src.common_utils.plot_error_distribution import write_images
from src.utils.model_utils.NewPreDeal_Tools import deal_verify_ydata, Deal_sorted_Ydata, deal_sorted_y_data

# warnings.filterwarnings(action='always', category=UserWarning)
warnings.filterwarnings(action='ignore')


# base_data_dirs = "../datas"
# 这个pandas处理数据效果不太好 建议用numpy


def gen_init_data(excel_path=None, excel_sheet=None):
    """
    复制一份去操作
    :param excel_path:
    :param excel_sheet:
    :return:
    """
    # 初始化
    excel_data_copy = copy.deepcopy(pd.read_excel(io=f'{excel_path}', sheet_name=excel_sheet))
    init_data = excel_data_copy
    # Preddata = pd.read_excel(io=r'data_original.xlsx', sheet_name='8000D数据标准化-预测用')
    # 用来查看数据结果的excel对象
    filename = f'{excel_path}'
    data = copy.deepcopy(load_workbook(f'{filename}'))
    # sheetnames = data.sheetnames
    # create_sheet
    # if len(sheetnames) < 4:
    #     sheetnames.append('预测结果')
    #     data.create_sheet(sheetnames[-1], len(sheetnames))
    #     # 赋值sheet
    #     # sheet=data[sheetnames[0]]
    #     # content=data.copy_worksheet(sheet)
    #     data.save(f'{filename}')
    # 预测结果的excel表
    sheetnames = data.sheetnames
    table = data[sheetnames[excel_sheet]]

    # 获取数据表对象(建模和与预测的数据)
    init_data = np.array(init_data)
    return init_data, table, filename, data


# 写入excel文件
def write_to_excel(start_index, table, current_column, content):
    # 标签
    for i in range(len(content)):
        table.cell(start_index, current_column).value = content[i]
        start_index += 1


def save_model(model, model_name, X_train, X_test, y_train, y_test, train_numbers):
    """
    :param model: 需要训练的模型
    :param X_train:
    :param X_test:
    :param y_train:
    :param y_test:
    :return:
    """
    # 训练次数
    if not os.path.exists(f"models/{model_name}"):
        os.makedirs(name=f"models/{model_name}", exist_ok=True)
    acc_dic = {}
    for i in range(11):
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        # 取最小的
        model_train_acc = min(model_train_acc, model_acc)
        acc_dic.update({f"model[{i}]": model_train_acc})
        # 模型保存
        joblib.dump(model, f'models/{model_name}/model[{i}].pkl')
    for i in range(11, train_numbers):
        # temp_keys 是键的按照值从小到打排序的列表 [model[1],model[5],...]
        temp_keys = sorted(acc_dic, reverse=True)
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        model_train_acc = min(model_train_acc, model_acc)
        # 替换一个最小的模型
        for key_item_index in range(0, len(temp_keys)):
            if key_item_index != len(temp_keys) - 1:
                if acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc < acc_dic[
                    f"{temp_keys[key_item_index + 1]}"]:
                    acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                    joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')
            elif acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc:
                acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')

    # model1 = joblib.load(filename="filename.pkl")


def develop_save_model(model=None, model_name=None, model_dict=None, save_path="./all_models",
                       x_train=None, y_train=None, train_numbers=1):
    """
    :param model: 需要训练的模型
    :param model_name:
    :param model_dict:
    :param save_path: 保存的路径，默认是当前文件夹下的 ./models
    :param x_train:
    :param y_train:
    :param train_numbers: 训练次数
    :return:
    """
    check_dir(save_path)
    # 只存储一个大模型
    model.fit(x_train, y_train.astype(float))
    model_dict.update({f'{model_name}': model})
    # 将字典中的所有模型保存到一个 joblib。可以考虑最后再保存，这样保证训练和预测可以同时进行
    joblib.dump(model_dict, f'{save_path}/models_temp.joblib')

    # for i in range(11):
    #     model.fit(X_train, y_train.astype(float))
    #     # 训练集准确率
    #     model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    #     # print('Rfc_train_acc准确率：', model_train_acc)
    #     model_pred = model.predict(X_test)
    #     model_acc = accuracy_score(y_test.astype(float), model_pred)
    #     # 取最小的
    #     model_train_acc = min(model_train_acc, model_acc)
    #     acc_dic.update({f"model[{i}]": model_train_acc})
    #     # 模型保存
    #     joblib.dump(model, f'models/{model_name}/model[{i}].pkl')
    # for i in range(11, train_numbers):
    #     # temp_keys 是键的按照值从小到打排序的列表 [model[1],model[5],...]
    #     temp_keys = sorted(acc_dic, reverse=True)
    #     model.fit(X_train, y_train.astype(float))
    #     # 训练集准确率
    #     model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    #     # print('Rfc_train_acc准确率：', model_train_acc)
    #     model_pred = model.predict(X_test)
    #     model_acc = accuracy_score(y_test.astype(float), model_pred)
    #     model_train_acc = min(model_train_acc, model_acc)
    #     # 替换一个最小的模型
    #     for key_item_index in range(0, len(temp_keys)):
    #         if key_item_index != len(temp_keys) - 1:
    #             if acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc < acc_dic[
    #                 f"{temp_keys[key_item_index + 1]}"]:
    #                 acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
    #                 joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')
    #         elif acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc:
    #             acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
    #             joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')


def check_dir(save_path):
    """ 检查目录是否存在，不存在就新建
    :param save_path:
    :return:
    """
    if not os.path.exists(f"{save_path}"):
        os.makedirs(name=f"{save_path}", exist_ok=True)


def get_pred_data(model_name, X_verify_data):
    """
    :param model_name: 模型名字
    :return: pred_data
    """
    # 预测数据
    all_models = os.listdir(f"models/{model_name}")
    model_pred_list = []
    for i in all_models:
        cursor_model = joblib.load(filename=f"models/{model_name}/{i}")
        model_pred_list.append(cursor_model.predict(X_verify_data))
    model_pred_list = np.array(model_pred_list)
    temp_pred_list = []
    # 循环遍历列
    for pred_list_index in range(0, model_pred_list.shape[1]):
        # 预测结果统一,那个数多就等于哪个
        zero_count = np.where(model_pred_list[:, pred_list_index] == 0)[0].shape[0]
        negative_count = np.where(model_pred_list[:, pred_list_index] == -1)[0].shape[0]
        positive_count = np.where(model_pred_list[:, pred_list_index] == 1)[0].shape[0]
        final_dic = {}
        final_dic.update({"0": zero_count})
        final_dic.update({"-1": negative_count})
        final_dic.update({"1": positive_count})
        # 排序，第一个数预测的最多
        final_result = -1
        if final_dic[f"{final_result}"] < final_dic["0"]:
            final_result = 0
        if final_dic[f"{final_result}"] < final_dic["1"]:
            final_result = 1
        temp_pred_list.append(final_result)
    pred_data = np.array(temp_pred_list)
    return pred_data


def get_train_test_acc(model, X_train, y_train, X_test, y_test):
    """
    :param model:
    :param X_train:
    :param y_train:
    :param X_test:
    :param y_test:
    :return: model_train_acc, model_acc
    """
    model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    model_pred = model.predict(X_test)
    model_acc = accuracy_score(y_test.astype(float), model_pred)
    return model_train_acc, model_acc


def get_develop_pred_data(excel_data=None, boundary_x=None, boundary_y=None):
    """
    :param excel_data:
    :param boundary_x:
    :param boundary_y:
    :return:
    """
    excel_data_copy = copy.deepcopy(excel_data)
    return excel_data_copy[boundary_x[0]:boundary_x[-1], boundary_y[0]:boundary_y[-1]]


def get_train_test_data(x_data=None, y_data=None):
    """ 按照奇偶分
    :param X_data:
    :param Y_data:
    :return: X_train, X_test, y_train, y_test
    """
    # 拷贝处理
    x_data_copy = copy.deepcopy(x_data)
    y_data_copy = copy.deepcopy(y_data)
    # 转化成ndarray操作
    X_train = []
    X_test = []
    Y_train = []
    Y_test = []
    iter_index = 1
    for i in range(x_data_copy.shape[0]):
        # 3 个一组 1 3 训练 2 测试
        if iter_index == 4:
            iter_index = 1
            X_train.append(x_data_copy[i, :])
            Y_train.append(y_data_copy[i])
        elif iter_index == 2:
            X_test.append(x_data_copy[i, :])
            Y_test.append(y_data_copy[i])
        else:
            X_train.append(x_data_copy[i, :])
            Y_train.append(y_data_copy[i])
        iter_index += 1
    X_train = np.array(X_train)
    X_test = np.array(X_test)
    Y_train = np.array(Y_train)
    Y_test = np.array(Y_test)
    return X_train, X_test, Y_train, Y_test


def get_usable_data(x_data=None, y_data=None):
    """
    获取删除无效行的数据,并且记录需要删除的行
    :param x_data: 删除无效列的
    :param y_data: 删除无效列的
    :return: 删除无效行以后的数据
    """
    # 先按照列合并
    temp_OriginalXdata = np.column_stack((x_data, y_data))
    # 删除缺失的行
    temp_OriginalXdata, del_raws = del_deletion_data(temp_OriginalXdata, 0)
    # X_data,Y_data 是处理好(行列均可用)的x,y
    X_data = temp_OriginalXdata[:, :-1]
    Y_data = temp_OriginalXdata[:, -1]
    print(
        f'空列表就是删除干净了.**{list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist()))}', )
    return X_data, Y_data, del_raws


def get_invert_y(y_develop_data=None, y_verify_data=None):
    """
    :param y_develop_data:建模的Y(源数据，未转化-1 0 1)
    :param y_verify_data: 预测的Y(源数据，未转化-1 0 1)
    :return: 转化好的 Y_develop_data,Verify_Ydata
    """
    # 获取 建模和预测转化好的Y
    Y_develop_data_copy = copy.deepcopy(y_develop_data)
    y_develop_data, Y_data_boundsMin, Y_data_boundsMax = Deal_sorted_Ydata(data=Y_develop_data_copy)
    # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
    # 保留原始的数据
    Y_verify_data_copy = copy.deepcopy(y_verify_data)
    # 验证（预测）的Y
    Verify_Ydata = deal_verify_ydata(Y_verify_data_copy, Y_data_boundsMin, Y_data_boundsMax)
    # 查看是否替换成功
    print("查看是否替换成功：", [unique(y_develop_data), Y_data_boundsMin, Y_data_boundsMax],
          np.unique(Verify_Ydata))
    return y_develop_data, Verify_Ydata





def develop_model(x_develop_data=None, x_pred_data=None, y_develop_data=None, y_pred_data=None, final_cols=None,
                  y_filename=None, traverse_index=None, start_month=None,
                  max_progress_row=20, data_interval=10,
                  the_column=None):
    """
    :param the_column: 这个是 out在excel的对应的列号
    :param x_develop_data:
    :param x_pred_data:
    :param y_develop_data:
    :param y_pred_data: 预测的y
    :param final_cols: 可用的列
    :param y_filename: 根据月份的y的文件位置
    :param traverse_index: 遍历次数，可用乘一个固定数，这样让excel的数据加进去
    :param start_month: 开始月份
    :param max_progress_row: 列的列表 最长 默认给了20
    :param data_interval: excel间隔，默认为10
    :return:
    """
    for index in range(len(final_cols)):
        # todo 后面想办法改一下，不要通过序列去传递了，很麻烦的！
        current_index = final_cols[index]
        y_develop_data_curr = y_develop_data[index]
        # y_verify_data_curr = y_pred_data[index]
        # 拷贝原数据，这个画图还是需要使用以前的数据
        Y_develop_data_orinal = copy.deepcopy(y_develop_data_curr)
        # Y_verify_data_orinal = copy.deepcopy(y_verify_data_curr)
        # 已经处理过列了，现在获取行可用的 建模数据,和已经删除的行
        X_develop_data, Y_develop_data, del_raws = get_usable_data(x_data=x_develop_data, y_data=y_develop_data_curr)
        # 处理验证（预测）的X_data,Y_data
        X_verify_data, Y_verify_data, del_raws = get_usable_data(x_data=x_pred_data, y_data=y_verify_data_curr)
        Y_develop_data, Verify_Ydata = get_invert_y(y_develop_data=Y_develop_data, y_verify_data=Y_verify_data)
        # 手动划分数据集，各自一半，按照奇偶
        X_train, X_test, y_train, y_test = get_train_test_data(x_data=X_develop_data, y_data=Y_develop_data)
        # ******************************************************
        # 降维，失败则返回原数据
        # todo 判断是否降维，后面再抽离出来
        """
        try:
            if int(X_train.shape[1]) > 7:
                # 降维
                II = train_model_ad(n_max=max(int(X_train.shape[1]/5),7), cd_n=r"./", cd_e="./", XT_SP_d_train=X_train,
                                    XT_SP_d_test=X_test, Y_train=y_train, Y_test=y_test,
                                    excel_name=f'{start_month + traverse_index}add_result')
                usable_cols = II[-1]
                X_train = X_train[:, usable_cols]
                X_test = X_test[:, usable_cols]
                # ########################################################
                # 验证集
                X_verify_data = X_verify_data[:, usable_cols]
        except Exception as e:
            print(f'降维失败，原因：{e}')
        """

        # ******************************************************
        # 验证集 这个可以不写，只是方便查看
        Y_verify_data = Verify_Ydata
        # 开始训练模型
        from sklearn.ensemble import RandomForestClassifier
        # 设置训练次数
        train_numbers = 15
        # 获取模型  随机森林
        Rfc = RandomForestClassifier()
        model_name = "Rfc"
        develop_save_model(Rfc, model_name, X_train, X_test, y_test)
        # 随便取一个作为训练集的结果
        current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
        Rfc_train_acc, Rfc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
        Rfc_pred = get_pred_data("Rfc", X_verify_data)
        verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)

        # todo 每个阶段不一样，后面这个会变动
        write_images(data_left=Y_develop_data_orinal, data_right=Y_verify_data_orinal,
                     names=f"{y_filename}/{all_titles[the_column + current_index - 1]}.png", score=verify_y1)
        # 计算准确率
        print('验证集Rfc_acc准确率:', verify_y1)
        #  标签 都加20
        write_to_excel(start_index=5 + traverse_index * data_interval, table=table, current_column=1,
                       content=[f"({start_month + traverse_index}月)训练集个数", "随机森林（Rfc）准确率",
                                "测试集个数", "随机森林（Rfc）准确率",
                                "验证集个数", "随机森林（Rfc）准确率", ])
        # 将训练集写入到excel中
        write_to_excel(start_index=5 + traverse_index * data_interval, table=table, current_column=the_column + index,
                       content=[X_train.shape[0], Rfc_train_acc,
                                X_test.shape[0], Rfc_acc,
                                X_verify_data.shape[0], verify_y1])
        # data_year_list = add_data_year(X_train_num=X_train.shape[0], Rfc_train_acc=Rfc_train_acc,
        #                                X_test_num=X_test.shape[0], Rfc_test_acc=Rfc_acc,
        #                                X_verify_data_num=X_verify_data.shape[0],
        #                                Rfc_verify_acc=verify_y1, data_year_list=data_year_list)
        # max_progress_row 表示当前最长
        # 达到一年会写入 目前是根据 traverse_index达到 12
        write_year_result(the_column=the_column, start_month=start_month, traverse_index=traverse_index, index=index,
                          data_interval=data_interval)
        data.save(filename)


def write_year_result(the_column=None, start_month=None, traverse_index=None, index=None,
                      data_interval=None, max_progress_row=20):
    """
    :param max_progress_row:
    :param the_column:
    :param start_month:
    :param traverse_index:
    :param index:
    :param data_interval:
    :return:
    """
    # 达到一年了，计算一下综合的数据，计算准确率
    if (traverse_index + 1) % 12 == 0:
        # todo 根据excel去操作
        # x_train_index=5 + traverse_index * data_interval
        # x_test_index=5 + traverse_index * data_interval
        # x_verify_index=5 + traverse_index * data_interval
        # [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
        #  x_verify_num_year] = data_year_list
        # x_train_acc_year = x_train_num_right_year / x_train_num_year
        # x_test_acc_year = x_test_num_right_year / x_test_num_year
        # x_verify_acc_year = x_verify_right_year / x_verify_num_year
        write_to_excel(start_index=5 + traverse_index * data_interval + data_interval, table=table, current_column=1,
                       content=[f"(从{start_month}月开始，一年)训练集个数", "随机森林（Rfc）准确率",
                                "测试集个数", "随机森林（Rfc）准确率",
                                "验证集个数", "随机森林（Rfc）准确率", ])
        # 将训练集写入到excel中
        write_to_excel(start_index=5 + traverse_index * data_interval + data_interval, table=table,
                       current_column=the_column + index,
                       content=["x_train_num_year", "x_train_acc_year",
                                "x_test_num_year", "x_test_acc_year",
                                "x_verify_num_year", "x_verify_acc_year"])


def add_data_year(X_train_num=None, Rfc_train_acc=None, X_test_num=None, Rfc_test_acc=None, X_verify_data_num=None,
                  Rfc_verify_acc=None, data_year_list=None):
    """
    这个方法没啥用了
    :param X_train_num:
    :param Rfc_train_acc:
    :param X_test_num:
    :param Rfc_test_acc:
    :param X_verify_data_num:
    :param Rfc_verify_acc:
    :param data_year_list:
    :return:
    """
    # 取出列表数据处理
    [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
     x_verify_num_year] = data_year_list
    # 训练个数
    x_train_num_right_year += X_train_num * Rfc_train_acc
    x_train_num_year += X_train_num
    # 测试
    x_test_num_right_year += X_test_num * Rfc_test_acc
    x_test_num_year += X_test_num
    # 验证
    x_verify_right_year += X_verify_data_num * Rfc_verify_acc
    x_verify_num_year += X_verify_data_num
    # 处理好数据再放回
    temp_list = [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year,
                 x_verify_right_year,
                 x_verify_num_year]
    return temp_list


def write_x_images(develop_x_title_data=None, develop_x_data=None, verify_x_data=None, x_filename="x_images"):
    """
    :param
    develop_x_title_data: 带标题的
    :param
    develop_x_data:
    :return:
    """
    # 将 建模的X 和 预测的X 传入即可，可先将标题带上
    title_data = list(develop_x_title_data)
    for i in range(len(develop_x_title_data)):
        write_images(data_left=develop_x_data[:, i], data_right=verify_x_data[:, i],
                     names=f"{x_filename}/{title_data[i]}.png")


def gen_data_year():
    """ 这个方法没啥用
    :return: data_year 训练，测试，验证的数据
    """
    x_train_num_right_year = 0
    x_train_num_year = 0
    x_test_num_right_year = 0
    x_test_num_year = 0
    x_verify_right_year = 0
    x_verify_num_year = 0
    data_year = [x_train_num_right_year, x_train_num_year, x_test_num_right_year, x_test_num_year, x_verify_right_year,
                 x_verify_num_year]
    return data_year


def convert_to_month(x_data=None, col=0):
    """
    :param x_data:
    :return: 将第一列(默认)转化为月份
    """
    x_data_copy = copy.deepcopy(x_data)
    first_row_data = list(x_data_copy[:, col])
    row_length = len(first_row_data)
    try:
        for index in range(row_length):
            if type(first_row_data[index]) == str:
                # 判断是否存在 “-”
                if "-" in first_row_data[index]:
                    first_row_data[index] = float(first_row_data[index].split("-")[-1])
                    # 取出最后两位作为月份
                    # first_row_data[index]=float(first_row_data[index][-2:])
    except Exception as e:
        print(e)
    # 取出最后两位作为月份
    # first_row_data = [float(item[-2:]) for item in first_row_data]
    first_row_data = np.array(first_row_data)
    x_data[:, col] = first_row_data
    return x_data


def get_boundary_list(excel_data=None):
    num = 2
    pattern = f"工序{num}-OUT"
    # todo
    pass
    # if re.match(pattern, cur_data):

    return list


def get_month_boundary_list(excel_data=None, start_year=None, start_month=None, pred_count=12, col=2):
    """
    :param excel_data:
    :param start_year:
    :param start_month:
    :param pred_count: 取多少个月的次数，因为中间有月份缺失
    :param col: 从excel第几列开始取数据
    :return:
    """
    start_year = int(start_year)
    start_month = int(start_month)
    # 当前匹配的年月份
    current_date = start_month
    # 是否匹配到开始月份
    start_flag = False
    # 定义正则
    pattern = r"\d{4}-\d{2}"
    # 将一列复制出来，转化为列表，并且转化为字符串
    month_list = copy.deepcopy(excel_data[:, col])
    # 去掉空格，并且转化为字符串
    month_list = [str(i) for i in month_list]
    boundary_list = []
    date_list = []
    for i in range(len(month_list)):
        cur_data = month_list[i]
        if start_flag:
            if pred_count > 0:
                if re.match(pattern, cur_data):
                    if cur_data != current_date:
                        current_date = cur_data
                        # 数据里面的数据有月份缺失
                        boundary_list.append(i)
                        date_list.append(current_date)
                        # 迭代次数减1
                        pred_count = pred_count - 1
            else:
                return boundary_list, date_list
        else:
            # 匹配 xxxx-xx
            if re.match(pattern, cur_data):
                # 匹配到开始月份
                if int(cur_data.split('-')[0]) == start_year and int(cur_data.split('-')[-1]) == start_month:
                    start_flag = True
                    current_date = cur_data
                    # 数据里面的数据有月份缺失
                    boundary_list.append(i)
                    date_list.append(current_date)
                    # 迭代次数减1
                    pred_count = pred_count - 1
    if len(boundary_list) == 0:
        raise Exception(f'{start_year}年的{start_month}月份不存在')
    return boundary_list, date_list


def del_deletion_data(data_value, flag):
    """
    :param data_value: 需要处理的数据
    :param flag: flag是1则删除列，如果是0则删除行
    :return: final_data, del_raws/del_cols
    """
    # 记录需要删除的行或者列
    del_list = []
    data_value_copy = copy.deepcopy(data_value)
    # 先拿到元数据的长度
    pd_data_value_copy = pd.DataFrame(copy.deepcopy(data_value_copy))
    # 获取所有存在的空值
    result_list = np.array(pd_data_value_copy.iloc[:, :].isnull().values.tolist())
    # 删除缺失值过多的列
    if flag == 1:
        # 计算多少列
        cols_length = result_list[0, :].size
        # 计算里面缺失值大于0.3的
        for index in range(cols_length):
            item = result_list[:, index]
            lock_list = [i for i in item if i]
            if len(lock_list) / len(item) > 0.3:
                del_list.append(index)
    # 删除含有缺失值的行
    elif flag == 0:
        rows_length = result_list[:, 0].size
        for index in range(rows_length):
            item = result_list[index, :]
            # 判断是否有缺失
            if True in item:
                del_list.append(index)
    np.delete(data_value_copy, del_list, axis=flag)
    return data_value_copy, del_list


def r_2(x, y):
    x_mean = np.mean(x)
    y_mean = np.mean(y)
    r2 = np.sum(np.multiply(x - x_mean, y - y_mean)) ** 2 / np.sum(np.power(x - x_mean, 2)) / np.sum(
        np.power(y - y_mean, 2))
    return r2


e = math.exp(1)


def DataDelete(X, r2_r):
    X = copy.deepcopy(X)
    try:
        X = np.array(X, dtype="float64")
        X = del_deletion_data(data_value=X, flag=0)
        X = np.matrix(X)
        nn = np.shape(X)[1]
        ii_l = np.arange(0, nn, 1)
        i = 0
        while i < nn - 1:
            for j in range(nn - 1, i, -1):
                r2 = r_2(X[:, i], X[:, j])
                if r2 > r2_r or np.isnan(np.sum(X[:, j])) or np.sum(X[:, j]) == 0:
                    ii_l = np.delete(ii_l, j, 0)
                    X = np.delete(X, j, 1)
            i = i + 1
            nn = np.shape(X)[1]
        return ii_l
    except Exception as e:
        print("异常：", e)
